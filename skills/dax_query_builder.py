"""
DAX Query Builder Module
========================
Reads PBI metadata extractor output (Excel) and generates DAX queries for each visual.

Input:  Metadata extractor Excel file (pbi_report_metadata_*.xlsx)
Output: Excel file with one row per visual containing the generated DAX query.

Usage:
    python dax_query_builder.py <input_excel> [output_excel]

Example:
    python dax_query_builder.py pbi_report_metadata_revopp.xlsx dax_queries_output.xlsx
"""

import sys
import os
import re
import argparse
from collections import OrderedDict, deque

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# =============================================================================
# CORE LOGIC: Field Classification
# =============================================================================

def classify_field(usage):
    """
    Classify a field's role in the DAX query based on its Usage label
    from the metadata extractor.

    Returns: 'grouping', 'measure', 'filter', 'slicer', or 'page_filter'
    """
    u = usage.lower()

    if "slicer" in u:
        return "slicer"
    if "page filter" in u:
        return "page_filter"

    # "Filter (Measure)" = measure dependency row from recursive resolution.
    # Must be checked BEFORE grouping roles, because usage like
    # "Visual Column, Filter (Measure)" contains "visual column" but is
    # actually a measure (or its dependency), not a grouping column.
    if "filter (measure)" in u:
        return "measure"

    # Matrix column-axis fields (must check before "visual column" since
    # "visual matrix column" contains "visual column")
    if "visual matrix column" in u:
        return "matrix_column"

    # Check for grouping roles (Visual Row, Visual Column, Visual Group)
    if any(k in u for k in ["visual row", "visual column", "visual group"]):
        return "grouping"

    # Check for measure roles (Visual Value, Visual X, Visual Y2, Visual Tooltip,
    # Visual Size, Visual Goal, Visual Trend, Visual Min, Visual Max, Visual Target)
    if any(k in u for k in ["visual value", "visual x", "visual y2", "visual tooltip",
                             "visual size", "visual goal", "visual trend",
                             "visual min", "visual max", "visual target"]):
        return "measure"

    # Check for filter (but not "Filter (Measure)" which was already handled above)
    if "filter" in u:
        return "filter"

    return "other"


def classify_visual_fields(fields):
    """
    Take a list of fields for a single visual and separate them into
    grouping columns, measures, and filters.

    Each field is a dict with keys: ui_name, usage, table_sm, col_sm
    """
    grouping = []
    measures = []
    filters = []
    slicer_fields = []
    matrix_columns = []

    for f in fields:
        role = classify_field(f["usage"])

        # Implicit measures (drag-and-drop aggregation) should be treated as measures
        # even if their usage label would normally classify them as grouping (e.g.
        # Table "Visual Column" with SUM aggregation)
        if f.get("agg_func") and role in ("grouping", "matrix_column", "other"):
            role = "measure"

        if role == "grouping":
            grouping.append(f)
        elif role == "measure":
            measures.append(f)
        elif role == "filter":
            filters.append(f)
        elif role == "slicer":
            slicer_fields.append(f)
        elif role == "page_filter":
            filters.append(f)
        elif role == "matrix_column":
            matrix_columns.append(f)

    return grouping, measures, filters, slicer_fields, matrix_columns


# =============================================================================
# CORE LOGIC: Implicit Measure Helpers
# =============================================================================

# Map PBI aggregation function names to DAX function names
AGG_FUNC_MAP = {
    "Sum": "SUM",
    "Avg": "AVERAGE",
    "Count": "COUNT",
    "Min": "MIN",
    "Max": "MAX",
    "CountNonNull": "COUNTA",
    "Median": "MEDIAN",
}

# DAX functions that require numeric columns — string columns need CONVERT wrapping
NUMERIC_ONLY_FUNCS = {"SUM", "AVERAGE", "MEDIAN"}


def _implicit_measure_dax(agg_func, table, column, model=None,
                          data_type="", model_source=""):
    """Generate DAX expression for an implicit measure (drag-and-drop aggregation).

    CONVERT logic (priority order):
    1. If model_source == "pbixray" → ALWAYS use SUMX+CONVERT for numeric funcs
       (pbixray reports ALL columns as string, types are unreliable)
    2. Else if data_type == "string" → use CONVERT (real PBIP says it's string)
    3. Else fallback to model.columns check (standalone CLI backward compat)
    4. Otherwise use plain aggregation function

    Args:
        agg_func: Aggregation function name (e.g. "Sum", "Avg")
        table: Table name in the semantic model
        column: Column name in the semantic model
        model: Optional SemanticModel for column type lookup
        data_type: Column data type from metadata Excel (optional)
        model_source: Semantic model source from metadata Excel (optional)

    Returns:
        DAX expression string, e.g. "SUM('Table'[Column])" or
        "SUMX('Table', CONVERT('Table'[Column], DOUBLE))" for string columns
    """
    dax_func = AGG_FUNC_MAP.get(agg_func, agg_func.upper())
    col_ref = f"'{table}'[{column}]"

    if dax_func in NUMERIC_ONLY_FUNCS:
        needs_convert = False

        if model_source == "pbixray":
            # pbixray marks ALL columns as string — always use CONVERT
            needs_convert = True
        elif data_type == "string":
            # Real PBIP data type says string — use CONVERT
            needs_convert = True
        elif model:
            # Fallback: check model.columns (standalone CLI without metadata columns)
            col_info = model.columns.get((table, column))
            if col_info and col_info.data_type == "string":
                needs_convert = True

        if needs_convert:
            x_func = dax_func + "X"
            return f"{x_func}('{table}', CONVERT({col_ref}, DOUBLE))"

    return f"{dax_func}({col_ref})"


def _measure_expression(m, model=None):
    """Return the DAX expression for a measure field.

    For implicit measures (drag-and-drop aggregation), returns e.g. SUM('Table'[Column]).
    For explicit measures, returns [MeasureName].
    """
    if m.get("agg_func"):
        return _implicit_measure_dax(
            m["agg_func"], m["table_sm"], m["measure_name"], model,
            data_type=m.get("data_type", ""),
            model_source=m.get("model_source", ""),
        )
    return f"[{m['measure_name']}]"


# =============================================================================
# CORE LOGIC: DAX Query Generation
# =============================================================================

def build_dax_query(grouping, measures, filters, slicer_fields, visual_type,
                    model=None, matrix_columns=None):
    """
    Build a DAX query string based on the classified fields.

    Args:
        model: Optional SemanticModel for column type lookup (implicit measure conversion)
        matrix_columns: Optional list of matrix column-axis fields

    Returns: (pattern_name, dax_query_string)
    """
    # Deduplicate measures by ui_name — dependency resolution in the metadata
    # extractor can create multiple rows per measure (one per source table).
    # For the kept row, extract the actual measure name from col_sm (first
    # comma-separated element), since ui_name is the display name which users
    # can rename in the visual (e.g. "Opportunity Revenue" vs actual measure "Revenue").
    seen_measures = set()
    unique_measures = []
    for m in measures:
        if m['ui_name'] not in seen_measures:
            seen_measures.add(m['ui_name'])
            m = dict(m)  # copy so we don't mutate the original
            m['measure_name'] = m['col_sm'].split(',')[0].strip()
            unique_measures.append(m)
    measures = unique_measures

    is_slicer = visual_type == "slicer"

    # ----- Pattern 2: Slicer (Columns Only) -----
    if is_slicer and slicer_fields:
        if len(slicer_fields) == 1:
            s = slicer_fields[0]
            dax = f"EVALUATE\nVALUES('{s['table_sm']}'[{s['col_sm']}])"
        else:
            cols = [f"'{s['table_sm']}'[{s['col_sm']}]" for s in slicer_fields]
            dax = "EVALUATE\nDISTINCT(\n    SELECTCOLUMNS(\n"
            select_parts = [f"        \"{s['ui_name']}\", '{s['table_sm']}'[{s['col_sm']}]" for s in slicer_fields]
            dax += ",\n".join(select_parts)
            dax += "\n    )\n)"
        return "Pattern 2: Columns Only", dax

    # ----- Pattern 1: Measures Only (Cards, KPIs) -----
    if not grouping and measures:
        if len(measures) == 1:
            m = measures[0]
            expr = _measure_expression(m, model)
            dax = f"EVALUATE\n{{ {expr} }}"
            pattern = "Pattern 1: Single Measure"
        else:
            pairs = [f"    \"{m['ui_name']}\", {_measure_expression(m, model)}" for m in measures]
            dax = "EVALUATE\nROW (\n" + ",\n".join(pairs) + "\n)"
            pattern = "Pattern 1: Multiple Measures"
        return pattern, dax

    # ----- Pattern 2: Columns Only (no measures) -----
    if grouping and not measures:
        if len(grouping) == 1:
            g = grouping[0]
            dax = f"EVALUATE\nVALUES('{g['table_sm']}'[{g['col_sm']}])"
        else:
            select_parts = [f"        \"{g['ui_name']}\", '{g['table_sm']}'[{g['col_sm']}]" for g in grouping]
            dax = "EVALUATE\nDISTINCT(\n    SELECTCOLUMNS(\n"
            dax += ",\n".join(select_parts)
            dax += "\n    )\n)"
        return "Pattern 2: Columns Only", dax

    # ----- Pattern 3M: Matrix Summary (row groupings + measures, no column-axis) -----
    if matrix_columns and grouping and measures:
        cols = [f"    '{g['table_sm']}'[{g['col_sm']}]" for g in grouping]
        pairs = [f"    \"{m['ui_name']}\", {_measure_expression(m, model)}" for m in measures]
        all_args = cols + pairs
        dax = "EVALUATE\nSUMMARIZECOLUMNS (\n" + ",\n".join(all_args) + "\n)"
        return "Pattern 3M: Matrix Summary", dax

    # ----- Pattern 3: Columns + Measures (Most Visuals) -----
    if grouping and measures:
        cols = [f"    '{g['table_sm']}'[{g['col_sm']}]" for g in grouping]
        pairs = [f"    \"{m['ui_name']}\", {_measure_expression(m, model)}" for m in measures]
        all_args = cols + pairs
        dax = "EVALUATE\nSUMMARIZECOLUMNS (\n" + ",\n".join(all_args) + "\n)"
        return "Pattern 3: Columns + Measures", dax

    return "Unknown", "-- Could not determine DAX pattern for this visual"


def build_matrix_values_query(matrix_columns):
    """Generate a preflight VALUES() query for matrix column-axis fields.

    Returns the DAX string that retrieves distinct values for the column-axis field(s).
    """
    if len(matrix_columns) == 1:
        mc = matrix_columns[0]
        return f"EVALUATE\nVALUES('{mc['table_sm']}'[{mc['col_sm']}])"
    # Multiple column-axis fields (rare) — return one query per field
    queries = []
    for mc in matrix_columns:
        queries.append(f"EVALUATE\nVALUES('{mc['table_sm']}'[{mc['col_sm']}])")
    return "\n\n".join(queries)


# =============================================================================
# FILTER LINEAGE: Auto-detect flat measures for Matrix pivot
# =============================================================================

def build_filter_graph(relationships) -> dict:
    """Build a directed filter-propagation graph from model relationships.

    In PBI, the "to" side (PK/dimension) filters the "from" side (FK/fact)
    by default. With crossFilteringBehavior=bothDirections, the reverse
    edge is added too. Inactive relationships are skipped.

    Args:
        relationships: list of TmdlRelationship objects

    Returns:
        dict mapping table name (str) → set of table names it can directly filter
    """
    graph = {}
    for rel in relationships:
        if not rel.is_active:
            continue

        # Default direction: to_table (dimension/PK) filters from_table (fact/FK)
        graph.setdefault(rel.to_table, set()).add(rel.from_table)

        # Bidirectional: add reverse edge
        if rel.cross_filtering == "bothDirections":
            graph.setdefault(rel.from_table, set()).add(rel.to_table)

    return graph


def can_filter_reach(graph, source_table, target_table) -> bool:
    """BFS through the filter graph to check if source_table can filter target_table.

    Args:
        graph: dict from build_filter_graph()
        source_table: table that provides the filter (e.g. column-axis table)
        target_table: table that needs to be reachable (e.g. measure's home table)

    Returns:
        True if target_table is reachable from source_table via filter propagation
    """
    if source_table == target_table:
        return True

    visited = set()
    queue = deque([source_table])
    visited.add(source_table)

    while queue:
        current = queue.popleft()
        for neighbor in graph.get(current, set()):
            if neighbor == target_table:
                return True
            if neighbor not in visited:
                visited.add(neighbor)
                queue.append(neighbor)

    return False


def auto_detect_flat_measures(measures, matrix_columns, model) -> list:
    """Detect measures that can't be filtered by the matrix column-axis table.

    Uses the semantic model's relationships to build a filter graph and checks
    whether each measure's home table is reachable from the column-axis table.
    Measures with no reachable path are returned as "flat" — they should be
    included in the pivot query without CALCULATE wrapping.

    Args:
        measures: list of deduplicated measure field dicts (with ui_name, table_sm)
        matrix_columns: list of matrix column-axis field dicts (with table_sm)
        model: SemanticModel with relationships

    Returns:
        list of measure ui_name strings that should be flat (unreachable)
    """
    if not model or not model.relationships or not matrix_columns:
        return []

    graph = build_filter_graph(model.relationships)
    column_axis_table = matrix_columns[0]["table_sm"]

    flat = []
    for m in measures:
        measure_table = m.get("table_sm", "")
        if not measure_table:
            continue
        if not can_filter_reach(graph, column_axis_table, measure_table):
            flat.append(m["ui_name"])

    return flat


def build_matrix_pivot_query(grouping, measures, matrix_columns, column_values,
                             model=None, flat_measures=None):
    """Generate a single pivoted CALCULATE query for a Matrix visual.

    For each pivot measure x each column value, creates a named CALCULATE expression
    that filters the column-axis field to that value. Flat measures (unrelated to the
    column-axis table) are included as-is without CALCULATE wrapping.

    When flat_measures is None and model has relationships, auto-detects which measures
    are unreachable from the column-axis table via filter lineage and treats them as flat.

    Args:
        grouping: List of row-axis grouping field dicts
        measures: List of measure field dicts
        matrix_columns: List of matrix column-axis field dicts
        column_values: List of distinct values for the column-axis field
        model: Optional SemanticModel for implicit measure resolution and lineage
        flat_measures: Optional list of measure ui_names that should NOT be pivoted
            (included once without CALCULATE wrapping). Use when a measure is unrelated
            to the column-axis table and returns BLANK when filtered by it.
            When None and model has relationships, auto-detection is used.
            Pass an explicit empty list [] to disable auto-detection and pivot all.

    Returns: (pattern_name, dax_query_string) or (pattern_name, dax_query_string, auto_flat)
        When auto-detection runs, returns a 3-tuple with the auto-detected flat measure names.
    """
    auto_flat = []

    # Deduplicate measures (same logic as build_dax_query)
    seen_measures = set()
    unique_measures = []
    for m in measures:
        if m['ui_name'] not in seen_measures:
            seen_measures.add(m['ui_name'])
            m = dict(m)
            m['measure_name'] = m['col_sm'].split(',')[0].strip()
            unique_measures.append(m)
    measures = unique_measures

    # Auto-detect flat measures via filter lineage when not explicitly provided
    if flat_measures is None and model and hasattr(model, 'relationships') and model.relationships:
        auto_flat = auto_detect_flat_measures(measures, matrix_columns, model)
        flat_names = set(auto_flat)
    else:
        flat_names = set(flat_measures or [])

    # Split into pivot vs flat
    pivot_measures = [m for m in measures if m['ui_name'] not in flat_names]
    flat_list = [m for m in measures if m['ui_name'] in flat_names]

    # Row grouping columns
    cols = [f"    '{g['table_sm']}'[{g['col_sm']}]" for g in grouping]

    # Build CALCULATE expressions: one per (column_value, pivot_measure) pair
    mc = matrix_columns[0]  # primary column-axis field
    mc_ref = f"'{mc['table_sm']}'[{mc['col_sm']}]"

    calc_parts = []
    for val in column_values:
        # Determine if value is numeric or string
        is_numeric = False
        try:
            float(val)
            is_numeric = True
        except (ValueError, TypeError):
            pass

        filter_val = str(val) if is_numeric else f"\"{val}\""

        for m in pivot_measures:
            expr = _measure_expression(m, model)
            label = f"{val} {m['ui_name']}"
            calc_parts.append(
                f"    \"{label}\", CALCULATE({expr}, {mc_ref} = {filter_val})"
            )

    # Flat measures: included once without CALCULATE
    flat_parts = [
        f"    \"{m['ui_name']}\", {_measure_expression(m, model)}" for m in flat_list
    ]

    all_args = cols + calc_parts + flat_parts
    dax = "EVALUATE\nSUMMARIZECOLUMNS (\n" + ",\n".join(all_args) + "\n)"
    if auto_flat:
        return "Pattern 3M: Matrix Pivot", dax, auto_flat
    return "Pattern 3M: Matrix Pivot", dax


def add_filter_comments(dax, filters):
    """Append filter comments to DAX query for unextracted filter values."""
    if filters:
        comments = [f"-- Filter: '{f['table_sm']}'[{f['col_sm']}] (value not extracted)" for f in filters]
        dax += "\n\n" + "\n".join(comments)
    return dax


def _is_measure_filter(expr: str) -> bool:
    """Detect if a DAX filter expression is measure-based (not column-based).

    Measure-based filters reference bare [MeasureName] without a 'Table' prefix,
    e.g. NOT ([Total Units YTD Var %] = BLANK()), [Revenue] > 1000, ISBLANK([KPI]).
    These CANNOT go inside CALCULATETABLE — they must wrap with FILTER() instead.

    Column-based filters use 'Table'[Column] syntax and ARE valid in CALCULATETABLE.
    """
    # Strip outer NOT/parentheses for analysis
    stripped = expr.strip()

    # Find all [Name] references in the expression
    # Bare measure refs: [Name] NOT preceded by ' (which would be 'Table'[Column])
    bare_refs = re.findall(r"(?<!')\[([^\]]+)\]", stripped)

    # Find all 'Table'[Column] qualified refs
    qualified_refs = re.findall(r"'[^']+'\[([^\]]+)\]", stripped)

    # If there are bare refs that aren't also qualified, it's a measure filter
    # Also check for BLANK(), ISBLANK patterns as strong signals
    has_blank = bool(re.search(r"BLANK\s*\(\s*\)|ISBLANK\s*\(", stripped, re.IGNORECASE))

    if bare_refs and not qualified_refs:
        # All refs are bare [Name] — measure filter
        return True
    if bare_refs and has_blank:
        # Has bare refs + BLANK check — measure filter
        return True

    return False


def wrap_dax_with_filters(base_dax: str, filter_exprs: list, pattern: str) -> str:
    """Wrap a base DAX query with bookmark filter expressions.

    For Pattern 1 (measures only), uses CALCULATE:
        EVALUATE { CALCULATE([Measure], filter1, filter2) }

    For all other patterns, uses CALCULATETABLE:
        EVALUATE CALCULATETABLE(<inner>, filter1, filter2)

    Measure-based filters (e.g. NOT ([Measure] = BLANK())) cannot go inside
    CALCULATETABLE — they are applied as an outer FILTER() wrapper instead.

    Args:
        base_dax: The original DAX query string (starts with EVALUATE)
        filter_exprs: List of DAX filter expression strings
        pattern: The DAX pattern name (e.g. "Pattern 1: Single Measure")

    Returns:
        Wrapped DAX query string
    """
    if not filter_exprs:
        return base_dax

    # Separate column filters (CALCULATETABLE-safe) from measure filters (need FILTER)
    column_filters = []
    measure_filters = []
    for expr in filter_exprs:
        if _is_measure_filter(expr):
            measure_filters.append(expr)
        else:
            column_filters.append(expr)

    # Strip any trailing filter comments from the base DAX
    lines = base_dax.split("\n")
    core_lines = []
    for line in lines:
        if line.strip().startswith("-- Filter:"):
            break
        core_lines.append(line)
    clean_dax = "\n".join(core_lines).rstrip()

    filter_args = ",\n    ".join(column_filters) if column_filters else ""

    # Pattern 1: Single Measure → CALCULATE
    if pattern == "Pattern 1: Single Measure":
        # Original: EVALUATE\n{ [Measure] } or EVALUATE\n{ SUM('Table'[Column]) }
        # Extract measure expression from { <expr> }
        measure_match = re.search(r"\{\s*(.+?)\s*\}", clean_dax)
        if measure_match:
            measure_ref = measure_match.group(1)
            # For single measure, all filters (column + measure) go in CALCULATE
            all_args = ",\n    ".join(filter_exprs)
            return (f"EVALUATE\n"
                    f"{{ CALCULATE({measure_ref},\n"
                    f"    {all_args}\n"
                    f") }}")

    # Pattern 1: Multiple Measures → CALCULATE for each measure in ROW
    if pattern == "Pattern 1: Multiple Measures":
        # Original: EVALUATE\nROW(\n    "Name", [Measure], ...\n)
        # Wrap each measure: "Name", CALCULATE([Measure], filters)
        # Simpler approach: wrap the whole ROW in CALCULATETABLE
        pass  # Fall through to CALCULATETABLE

    # All other patterns: CALCULATETABLE (column filters) + FILTER (measure filters)
    # Strip the leading "EVALUATE\n" to get the inner expression
    if clean_dax.upper().startswith("EVALUATE"):
        inner = clean_dax[len("EVALUATE"):].strip()
    else:
        inner = clean_dax

    # Build CALCULATETABLE with column filters only
    if column_filters:
        result = (f"CALCULATETABLE(\n"
                  f"    {inner},\n"
                  f"    {filter_args}\n"
                  f")")
    else:
        result = inner

    # Wrap with FILTER for measure-based filters
    if measure_filters:
        condition = " && ".join(measure_filters)
        result = (f"FILTER(\n"
                  f"    {result},\n"
                  f"    {condition}\n"
                  f")")

    return f"EVALUATE\n{result}"


def wrap_dax_with_having(dax: str, having_exprs: list) -> str:
    """Wrap a DAX query with post-aggregation FILTER conditions.

    Wraps the entire query result in FILTER() to keep only rows matching
    the condition. Used for thresholds like "TotalSales > 1000000".

    Args:
        dax: The DAX query string (may already be CALCULATETABLE-wrapped)
        having_exprs: List of DAX boolean expressions (e.g., "[TotalSales] > 1000000")

    Returns:
        Wrapped DAX query string
    """
    if not having_exprs:
        return dax

    # Strip any trailing filter comments
    lines = dax.split("\n")
    core_lines = []
    for line in lines:
        if line.strip().startswith("-- Filter:"):
            break
        core_lines.append(line)
    clean_dax = "\n".join(core_lines).rstrip()

    # Strip the leading EVALUATE to get the inner expression
    if clean_dax.upper().startswith("EVALUATE"):
        inner = clean_dax[len("EVALUATE"):].strip()
    else:
        inner = clean_dax

    # Build combined condition with &&
    condition = " && ".join(having_exprs)

    return (f"EVALUATE\n"
            f"FILTER(\n"
            f"    {inner},\n"
            f"    {condition}\n"
            f")")


def parse_filter_column_refs(filter_exprs):
    """Extract (table, column) pairs from DAX filter expressions.

    Parses expressions like "'Opportunities'[Status] IN {\"Open\"}" and
    returns [("Opportunities", "Status")].
    """
    refs = []
    # Match 'TableName'[ColumnName] in filter expressions
    pattern = re.compile(r"'([^']+)'\[([^\]]+)\]")
    for expr in filter_exprs:
        for match in pattern.finditer(expr):
            ref = (match.group(1).strip(), match.group(2).strip())
            if ref not in refs:
                refs.append(ref)
    return refs


def check_filter_redundancy(measures, filter_exprs, model=None):
    """Check if filter expressions target columns already referenced in measure formulas.

    For each measure in the visual, looks up its DAX formula and checks whether
    the formula already references the same Table[Column] being filtered externally.
    This detects cases where CALCULATETABLE would conflict with internal measure logic.

    Args:
        measures: List of field dicts (must have col_sm, table_sm, measure_formula keys)
        filter_exprs: List of DAX filter expression strings
        model: Optional SemanticModel for fallback formula lookup

    Returns:
        List of warning dicts with keys: measure_name, filter_expr, filter_table,
        filter_column, measure_formula
    """
    filter_refs = parse_filter_column_refs(filter_exprs)
    if not filter_refs:
        return []

    warnings = []
    for m in measures:
        formula = m.get("measure_formula", "")
        # Fallback: look up formula from semantic model if not in metadata
        if not formula and model:
            formula = model.measures.get((m.get("table_sm", ""), m.get("col_sm", "")), "")
        if not formula:
            continue

        formula_upper = formula.upper()
        for (ftable, fcol) in filter_refs:
            # Check 'Table'[Column], Table[Column], and bare [Column] (same-table)
            if (f"'{ftable}'[{fcol}]".upper() in formula_upper or
                f"{ftable}[{fcol}]".upper() in formula_upper or
                (f"[{fcol}]".upper() in formula_upper and
                 m.get("table_sm", "").upper() == ftable.upper())):
                warnings.append({
                    "measure_name": m.get("col_sm", ""),
                    "filter_expr": next((e for e in filter_exprs if f"'{ftable}'[{fcol}]" in e), filter_exprs[0]),
                    "filter_table": ftable,
                    "filter_column": fcol,
                    "measure_formula": formula[:80],
                })
    return warnings


def build_bookmark_queries(bookmarks, visuals, page_filters, model=None):
    """Build bookmark-aware DAX queries for all visible visuals in each bookmark.

    Args:
        bookmarks: List of bookmark dicts from read_extractor_output
        visuals: OrderedDict of (page, visual_name) → visual data
        page_filters: Dict of page_name → filter fields
        model: Optional SemanticModel for fallback measure formula lookup

    Returns:
        List of dicts, each representing one row in the Bookmark DAX Queries sheet
    """
    if not bookmarks:
        return []

    # Group bookmark rows by bookmark name to get filter + visibility info
    from collections import defaultdict
    bm_groups = defaultdict(lambda: {"filter_dax": "", "page_name": "", "visuals": {}})

    for bm_row in bookmarks:
        bm_name = bm_row["bookmark_name"]
        bm_groups[bm_name]["page_name"] = bm_row["page_name"]
        bm_groups[bm_name]["filter_dax"] = bm_row["filter_dax"]
        bm_groups[bm_name]["visuals"][bm_row["visual_name"]] = bm_row["visible"]

    results = []

    for bm_name, bm_data in bm_groups.items():
        page_name = bm_data["page_name"]
        filter_dax_str = bm_data["filter_dax"]
        filter_exprs = [f.strip() for f in filter_dax_str.split(";") if f.strip()] if filter_dax_str else []

        # For each visible visual in this bookmark, find its base DAX query
        for visual_name, visible in bm_data["visuals"].items():
            if visible != "Y":
                continue

            # Find the visual by matching visual_name in data dicts
            # (keys may use visual_id instead of visual_name)
            matching_key = None
            for vkey, vdata in visuals.items():
                if vkey[0] == page_name and vdata["visual_name"] == visual_name:
                    matching_key = vkey
                    break
            if matching_key is None:
                continue

            data = visuals[matching_key]
            visual_type = data["visual_type"]
            fields = data["fields"]

            # Classify fields and build base DAX
            grouping, measures, filters, slicer_fields, matrix_columns = classify_visual_fields(fields)
            if page_name in page_filters:
                filters.extend(page_filters[page_name])

            pattern, base_dax = build_dax_query(grouping, measures, filters, slicer_fields,
                                                visual_type, model, matrix_columns=matrix_columns)

            if pattern == "Unknown":
                continue

            # Check for filter redundancy before wrapping
            active_filters = list(filter_exprs)
            if active_filters:
                measure_fields = [f for f in fields if classify_field(f["usage"]) == "measure"]
                redundancy_warnings = check_filter_redundancy(measure_fields, active_filters, model)
                if redundancy_warnings:
                    conflicting = {w["filter_expr"] for w in redundancy_warnings}
                    for w in redundancy_warnings:
                        print(f"WARNING [{bm_name} / {visual_name}]: "
                              f"Filter '{w['filter_expr']}' targets "
                              f"'{w['filter_table']}'[{w['filter_column']}] which is "
                              f"already referenced in [{w['measure_name']}]")
                        print(f"  Formula: {w['measure_formula']}...")
                        print(f"  Skipping this filter to avoid result mismatch.")
                    active_filters = [f for f in active_filters if f not in conflicting]

            # Wrap with bookmark filters
            if active_filters:
                wrapped_dax = wrap_dax_with_filters(base_dax, active_filters, pattern)
            else:
                wrapped_dax = base_dax

            filters_applied_str = "; ".join(active_filters) if active_filters else ""
            skipped_note = ""
            if len(active_filters) < len(filter_exprs):
                skipped = [f for f in filter_exprs if f not in active_filters]
                skipped_note = " | SKIPPED (redundant): " + "; ".join(skipped)

            results.append({
                "bookmark_name": bm_name,
                "page_name": page_name,
                "visual_name": visual_name,
                "visual_type": visual_type,
                "dax_pattern": pattern,
                "dax_query": wrapped_dax,
                "filters_applied": filters_applied_str + skipped_note,
                "validated": "",
            })

    return results


# =============================================================================
# INPUT: Read Metadata Extractor Excel
# =============================================================================

def read_extractor_output(filepath):
    """
    Read the metadata extractor Excel and return structured data.

    Returns:
        visuals: OrderedDict keyed by (page, visual_id_or_name) -> {
            'visual_type': str,
            'visual_name': str,
            'visual_id': str,
            'fields': [{'ui_name', 'usage', 'table_sm', 'col_sm'}, ...]
        }
        page_filters: dict keyed by page_name -> [field_dicts]
        bookmarks: list of dicts with bookmark data (empty if no Bookmarks sheet)
        filter_expr_data: list of dicts from Filter Expressions sheet (empty if absent)
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    col_map = {h: i for i, h in enumerate(headers)}

    # "Measure Formula" is optional — present when Skill 1 writes it, absent in older/manual files
    formula_idx = col_map.get("Measure Formula")

    # "Visual ID" is optional — present in new outputs, absent in older/manual files
    visual_id_idx = col_map.get("Visual ID")

    # "Aggregation Function" is optional — present when implicit measures exist
    agg_func_idx = col_map.get("Aggregation Function")
    # "Data Type" and "Semantic Model Source" — optional, added for pbixray workaround
    data_type_idx = col_map.get("Data Type")
    model_source_idx = col_map.get("Semantic Model Source")
    has_visual_id = visual_id_idx is not None

    required = ["Page Name", "Visual/Table Name in PBI", "Visual Type",
                 "UI Field Name", "Usage (Visual/Filter/Slicer)",
                 "Table in the Semantic Model", "Column in the Semantic Model"]

    for r in required:
        if r not in col_map:
            print(f"Error: Missing required column '{r}' in the input Excel.")
            print(f"Available columns: {headers}")
            sys.exit(1)

    visuals = OrderedDict()
    page_filters = {}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        page = row[col_map["Page Name"]]
        visual_name = row[col_map["Visual/Table Name in PBI"]]
        visual_type = row[col_map["Visual Type"]]
        ui_name = row[col_map["UI Field Name"]]
        usage = row[col_map["Usage (Visual/Filter/Slicer)"]]
        table_sm = row[col_map["Table in the Semantic Model"]]
        col_sm = row[col_map["Column in the Semantic Model"]]
        visual_id = (row[visual_id_idx] if has_visual_id else "") or ""

        if not page or not visual_name:
            continue

        field = {
            "ui_name": ui_name or "",
            "usage": usage or "",
            "table_sm": table_sm or "",
            "col_sm": col_sm or "",
            "measure_formula": (row[formula_idx] if formula_idx is not None else "") or "",
            "agg_func": (row[agg_func_idx] if agg_func_idx is not None else "") or "",
            "data_type": (row[data_type_idx] if data_type_idx is not None else "") or "",
            "model_source": (row[model_source_idx] if model_source_idx is not None else "") or "",
        }

        # Separate page-level filters
        if visual_type == "pageFilter":
            if page not in page_filters:
                page_filters[page] = []
            page_filters[page].append(field)
            continue

        # Use visual_id for grouping when available, fall back to visual_name
        if has_visual_id and visual_id:
            key = (page, visual_id)
        else:
            key = (page, visual_name)

        if key not in visuals:
            visuals[key] = {
                "visual_type": visual_type,
                "visual_name": visual_name,
                "visual_id": visual_id,
                "fields": [],
            }
        visuals[key]["fields"].append(field)

    # Read Bookmarks sheet if present
    bookmarks = []
    if "Bookmarks" in wb.sheetnames:
        ws_bm = wb["Bookmarks"]
        bm_headers = [cell.value for cell in ws_bm[1]]
        bm_col_map = {h: i for i, h in enumerate(bm_headers)}

        for row in ws_bm.iter_rows(min_row=2, max_row=ws_bm.max_row, values_only=True):
            if not row or not row[0]:
                continue
            bookmarks.append({
                "bookmark_name": row[bm_col_map.get("Bookmark Name", 0)] or "",
                "page_name": row[bm_col_map.get("Page Name", 1)] or "",
                "container_id": row[bm_col_map.get("Visual Container ID", 2)] or "",
                "visual_name": row[bm_col_map.get("Visual Name", 3)] or "",
                "visible": row[bm_col_map.get("Visible", 4)] or "",
                "filter_dax": row[bm_col_map.get("Filter DAX", 5)] or "",
            })

    # Read Filter Expressions sheet if present
    filter_expr_data = []
    if "Filter Expressions" in wb.sheetnames:
        ws_fe = wb["Filter Expressions"]
        fe_headers = [cell.value for cell in ws_fe[1]]
        fe_col_map = {h: i for i, h in enumerate(fe_headers)}

        for row in ws_fe.iter_rows(min_row=2, max_row=ws_fe.max_row, values_only=True):
            if not row or not row[0]:
                continue
            filter_expr_data.append({
                "page_name": row[fe_col_map.get("Page Name", 0)] or "",
                "visual_name": row[fe_col_map.get("Visual Name", 1)] or "",
                "visual_id": row[fe_col_map.get("Visual ID", 2)] or "",
                "filter_level": row[fe_col_map.get("Filter Level", 3)] or "",
                "filter_field": row[fe_col_map.get("Filter Field", 4)] or "",
                "filter_dax_expr": row[fe_col_map.get("Filter DAX Expression", 5)] or "",
            })

    wb.close()
    return visuals, page_filters, bookmarks, filter_expr_data


# =============================================================================
# FILTER EXPRESSION HELPERS
# =============================================================================

def collect_filters_for_visual(page_name, visual_name, visual_id, filter_expr_data):
    """Collect applicable filter DAX expressions following hierarchy: Report → Page → Slicer → Visual.

    Skips TopN and unsupported entries (starting with --).

    Args:
        page_name: Page display name
        visual_name: Visual display name
        visual_id: Visual container ID
        filter_expr_data: List of dicts from Filter Expressions sheet

    Returns:
        List of DAX filter expression strings applicable to this visual.
    """
    filters = []
    for fe in filter_expr_data:
        level = fe["filter_level"]
        dax_expr = fe["filter_dax_expr"]

        # Skip unsupported/comment entries
        if not dax_expr or dax_expr.startswith("--"):
            continue

        # Report-level: applies to all visuals
        if level == "Report":
            filters.append(dax_expr)
        # Page-level: applies to all visuals on that page
        elif level == "Page" and fe["page_name"] == page_name:
            filters.append(dax_expr)
        # Visual-level: applies to this specific visual only
        elif level == "Visual" and fe["page_name"] == page_name:
            # Match by visual_id first (preferred), fall back to visual_name
            # Page check is required because visual container IDs can collide across pages
            if visual_id and fe["visual_id"] == visual_id:
                filters.append(dax_expr)
            elif not visual_id and fe["visual_name"] == visual_name:
                filters.append(dax_expr)
        # Slicer-level: persisted slicer selection applies to all visuals on the page
        # (except the slicer itself)
        elif level == "Slicer" and fe["page_name"] == page_name:
            # Don't apply slicer filter back to the slicer visual itself
            if visual_id and fe["visual_id"] == visual_id:
                continue
            if not visual_id and fe["visual_name"] == visual_name:
                continue
            filters.append(dax_expr)

    return filters


def _is_measure_filter(dax_expr):
    """Detect whether a DAX filter expression is post-aggregation (needs FILTER wrapping).

    Post-aggregation filters:
      - Aggregation-wrapped columns: MIN('T'[C]) > DATE(...)  → True
      - Bare measure references:     [Rev Goal] > 0           → True

    Pre-aggregation (column) filters:
      - Direct column comparisons:   'T'[Status] IN {"Open"}  → False

    Examples:
        "'Opportunities'[Status] IN {\"Open\"}"              → False (column filter)
        "[Rev Goal] > 0"                                     → True  (measure filter)
        "MIN('Nations WH_Claims'[loss_date]) > DATE(2020,1,1)" → True  (aggregation filter)
    """
    # Aggregation function wrapping a column → post-aggregation filter
    if re.search(r'\b(SUM|AVERAGE|COUNT|COUNTA|MIN|MAX|MEDIAN|SUMX|AVERAGEX|COUNTX|MINX|MAXX)\s*\(', dax_expr):
        return True
    # Column filter: contains 'Table'[Column] pattern (no aggregation wrapper)
    if re.search(r"'[^']+'\[[^\]]+\]", dax_expr):
        return False
    # Bare [Measure] reference without table prefix
    if re.search(r"(?<!')\[[^\]]+\]", dax_expr):
        return True
    return False


# =============================================================================
# OUTPUT: Write DAX Queries to Excel
# =============================================================================

def write_output(visuals, page_filters, output_path, bookmark_queries=None,
                 filter_expr_data=None, model=None):
    """Write the DAX queries to a formatted Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DAX Queries by Visual"

    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F3864")
    normal_font = Font(name="Calibri", size=10, color="333333")
    code_font = Font(name="Consolas", size=9, color="333333")
    code_fill = PatternFill("solid", fgColor="F5F5F5")
    alt_fill = PatternFill("solid", fgColor="F2F2F2")
    wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center = Alignment(horizontal="center", vertical="top", wrap_text=True)
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    headers = [
        "Page Name",
        "Visual Name",
        "Visual Type",
        "DAX Pattern",
        "DAX Query",
        "Filtered DAX Query",
        "Filter Fields",
        "Matrix Column Field",
        "Preflight VALUES Query",
        "Validated?"
    ]
    col_widths = [22, 32, 22, 24, 70, 70, 30, 30, 50, 12]

    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap
        cell.border = thin

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    filter_expr_data = filter_expr_data or []

    # Process each visual
    row_num = 2
    for idx, ((page, _key), data) in enumerate(visuals.items()):
        visual_name = data["visual_name"]
        visual_id = data.get("visual_id", "")
        visual_type = data["visual_type"]
        fields = data["fields"]

        # Classify fields
        grouping, measures, filters, slicer_fields, matrix_columns = classify_visual_fields(fields)

        # Add page filters to filter list
        if page in page_filters:
            filters.extend(page_filters[page])

        # Build base DAX query
        pattern, dax = build_dax_query(grouping, measures, filters, slicer_fields,
                                       visual_type, model, matrix_columns=matrix_columns)

        # Add filter comments
        dax = add_filter_comments(dax, filters)

        # Build filtered DAX query (if filter expressions available)
        filtered_dax = ""
        if filter_expr_data and pattern != "Unknown":
            applicable_filters = collect_filters_for_visual(
                page, visual_name, visual_id, filter_expr_data,
            )
            if applicable_filters:
                # Run redundancy check against measure formulas
                measure_fields = [f for f in fields if classify_field(f["usage"]) == "measure"]
                redundancy_warnings = check_filter_redundancy(measure_fields, applicable_filters, model)
                active_filters = list(applicable_filters)
                if redundancy_warnings:
                    conflicting = {w["filter_expr"] for w in redundancy_warnings}
                    for w in redundancy_warnings:
                        print(f"WARNING [{visual_name}]: Filter '{w['filter_expr']}' targets "
                              f"'{w['filter_table']}'[{w['filter_column']}] — already in "
                              f"[{w['measure_name']}], skipping.")
                    active_filters = [f for f in active_filters if f not in conflicting]

                # Separate column filters from measure filters
                col_filters = [f for f in active_filters if not _is_measure_filter(f)]
                meas_filters = [f for f in active_filters if _is_measure_filter(f)]

                # Start with base DAX (without filter comments)
                clean_base = dax.split("\n\n-- Filter:")[0]
                filtered = clean_base

                # Apply column filters via CALCULATETABLE
                if col_filters:
                    filtered = wrap_dax_with_filters(filtered, col_filters, pattern)

                # Apply measure filters via FILTER (post-aggregation)
                if meas_filters:
                    filtered = wrap_dax_with_having(filtered, meas_filters)

                filtered_dax = filtered

        # Format filter field names
        filter_str = ", ".join([f"'{f['table_sm']}'[{f['col_sm']}]" for f in filters]) if filters else "None"

        # Matrix column-axis info
        matrix_col_str = ""
        values_query_str = ""
        if matrix_columns and pattern.startswith("Pattern 3M"):
            matrix_col_str = ", ".join(
                [f"'{mc['table_sm']}'[{mc['col_sm']}]" for mc in matrix_columns]
            )
            values_query_str = build_matrix_values_query(matrix_columns)

        # Write row
        row_data = [page, visual_name, visual_type, pattern, dax, filtered_dax,
                    filter_str, matrix_col_str, values_query_str, ""]
        is_alt = idx % 2 == 1

        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=j, value=val)
            cell.border = thin
            cell.alignment = wrap
            if j in (5, 6, 9):  # DAX query columns + preflight query
                cell.font = code_font
                cell.fill = code_fill
            else:
                cell.font = normal_font
                if is_alt:
                    cell.fill = alt_fill

        ws.row_dimensions[row_num].height = 80
        row_num += 1

    # Freeze and filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{row_num - 1}"

    # --- Bookmark DAX Queries sheet ---
    if bookmark_queries:
        ws_bm = wb.create_sheet("Bookmark DAX Queries")

        bm_headers = [
            "Bookmark Name",
            "Page Name",
            "Visual Name",
            "Visual Type",
            "DAX Pattern",
            "DAX Query",
            "Filters Applied",
            "Validated?"
        ]
        bm_col_widths = [24, 22, 32, 22, 24, 70, 40, 12]

        for i, h in enumerate(bm_headers, 1):
            cell = ws_bm.cell(row=1, column=i, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = wrap
            cell.border = thin

        for i, w in enumerate(bm_col_widths, 1):
            ws_bm.column_dimensions[get_column_letter(i)].width = w

        bm_row_num = 2
        for idx, bq in enumerate(bookmark_queries):
            row_data = [
                bq["bookmark_name"],
                bq["page_name"],
                bq["visual_name"],
                bq["visual_type"],
                bq["dax_pattern"],
                bq["dax_query"],
                bq["filters_applied"],
                bq["validated"],
            ]
            is_alt = idx % 2 == 1

            for j, val in enumerate(row_data, 1):
                cell = ws_bm.cell(row=bm_row_num, column=j, value=val)
                cell.border = thin
                cell.alignment = wrap
                if j == 6:  # DAX query column
                    cell.font = code_font
                    cell.fill = code_fill
                else:
                    cell.font = normal_font
                    if is_alt:
                        cell.fill = alt_fill

            ws_bm.row_dimensions[bm_row_num].height = 100
            bm_row_num += 1

        ws_bm.freeze_panes = "A2"
        ws_bm.auto_filter.ref = f"A1:H{bm_row_num - 1}"
        print(f"Generated bookmark DAX queries for {len(bookmark_queries)} visual×bookmark combinations")

    wb.save(output_path)
    return row_num - 2  # number of visuals processed


# =============================================================================
# SINGLE VISUAL QUERY: Look up one visual and optionally wrap with filters
# =============================================================================

def find_visual(visuals, search_term):
    """Find a visual by name (case-insensitive, partial match).

    Matches against visual_name (from data dict) and against "page / visual" combined.
    Returns: list of keys that match, best matches first.
    """
    search_lower = search_term.lower()
    exact = []
    partial = []

    for key, data in visuals.items():
        page = key[0]
        visual_name = data["visual_name"]
        name_lower = visual_name.lower()
        full_lower = f"{page} / {visual_name}".lower()
        if name_lower == search_lower or full_lower == search_lower:
            exact.append(key)
        elif search_lower in name_lower or search_lower in full_lower:
            partial.append(key)

    return exact + partial


def get_single_visual_query(visuals, page_filters, visual_search,
                            filter_exprs=None, having_exprs=None, model=None,
                            filter_expr_data=None, column_values=None,
                            flat_measures=None):
    """Look up a single visual's DAX query and optionally wrap with filters.

    Automatically applies preset filters from the Filter Expressions sheet
    (report-level, page-level, visual-level) before applying any explicit
    filter_exprs passed by the caller. Explicit filters are additive.

    For Matrix visuals with column-axis fields (Pattern 3M), returns the
    preflight VALUES() query and matrix column info. If column_values is
    provided, generates the pivoted CALCULATE query.

    Args:
        visuals: OrderedDict from read_extractor_output
        page_filters: dict from read_extractor_output
        visual_search: visual name search string (partial match OK)
        filter_exprs: optional list of pre-aggregation DAX filter expressions (CALCULATETABLE)
        having_exprs: optional list of post-aggregation DAX conditions (FILTER)
        model: optional SemanticModel for fallback measure formula lookup
        filter_expr_data: optional list of dicts from Filter Expressions sheet —
            preset report/page/visual filters are auto-applied when present
        column_values: optional list of distinct values for matrix column-axis field —
            when provided, generates the pivoted CALCULATE query
        flat_measures: optional list of measure ui_names that should NOT be pivoted —
            included once without CALCULATE wrapping (for measures unrelated to column-axis)

    Returns:
        dict with keys: page, visual_name, visual_type, pattern, dax_query,
        base_dax_query, filters_applied, having_applied, preset_filters_applied,
        matrix_columns, values_query, pivot_dax_query
        or None if no match found.
    """
    matches = find_visual(visuals, visual_search)

    if not matches:
        print(f"No visual found matching '{visual_search}'")
        print(f"\nAvailable visuals:")
        for key, vdata in visuals.items():
            print(f"  {key[0]} / {vdata['visual_name']}")
        return None

    if len(matches) > 1:
        print(f"Multiple visuals match '{visual_search}':")
        for i, key in enumerate(matches, 1):
            print(f"  {i}. {key[0]} / {visuals[key]['visual_name']}")
        print(f"\nUsing first match: {visuals[matches[0]]['visual_name']}")

    key = matches[0]
    page = key[0]
    data = visuals[key]
    visual_name = data["visual_name"]
    visual_id = data.get("visual_id", "")
    visual_type = data["visual_type"]
    fields = data["fields"]

    # Classify fields and build base DAX
    grouping, measures, filters, slicer_fields, matrix_columns = classify_visual_fields(fields)
    if page in page_filters:
        filters.extend(page_filters[page])

    pattern, base_dax = build_dax_query(grouping, measures, filters, slicer_fields,
                                        visual_type, model, matrix_columns=matrix_columns)

    # --- Matrix pivot support ---
    values_query = ""
    pivot_dax_query = ""
    matrix_col_info = []
    auto_flat_measures = []
    if matrix_columns and pattern.startswith("Pattern 3M"):
        values_query = build_matrix_values_query(matrix_columns)
        matrix_col_info = [
            {"table": mc["table_sm"], "column": mc["col_sm"], "ui_name": mc["ui_name"]}
            for mc in matrix_columns
        ]
        if column_values:
            pivot_result = build_matrix_pivot_query(
                grouping, measures, matrix_columns, column_values, model,
                flat_measures=flat_measures
            )
            # 3-tuple when auto-detection found flat measures, 2-tuple otherwise
            if len(pivot_result) == 3:
                _, pivot_dax_query, auto_flat_measures = pivot_result
            else:
                _, pivot_dax_query = pivot_result

    # --- Auto-collect preset filters from Filter Expressions sheet ---
    preset_filters = []
    if filter_expr_data:
        preset_filters = collect_filters_for_visual(
            page, visual_name, visual_id, filter_expr_data,
        )

    # Merge preset filters with explicit caller-provided filters (additive)
    all_filters = list(preset_filters)
    if filter_exprs:
        # Avoid duplicates — only add explicit filters not already in preset
        existing = set(all_filters)
        for f in filter_exprs:
            if f not in existing:
                all_filters.append(f)

    # Check for filter redundancy before wrapping with CALCULATETABLE
    active_filters = list(all_filters)
    if active_filters:
        measure_fields = [f for f in fields if classify_field(f["usage"]) == "measure"]
        redundancy_warnings = check_filter_redundancy(measure_fields, active_filters, model)
        if redundancy_warnings:
            conflicting = {w["filter_expr"] for w in redundancy_warnings}
            for w in redundancy_warnings:
                print(f"WARNING: Filter '{w['filter_expr']}' targets "
                      f"'{w['filter_table']}'[{w['filter_column']}] which is "
                      f"already referenced in [{w['measure_name']}]")
                print(f"  Formula: {w['measure_formula']}...")
                print(f"  Skipping this filter to avoid result mismatch.")
            active_filters = [f for f in active_filters if f not in conflicting]

    # Separate column filters from measure filters
    col_filters = [f for f in active_filters if not _is_measure_filter(f)]
    meas_filters = [f for f in active_filters if _is_measure_filter(f)]

    # Build filtered DAX query
    filtered_dax = base_dax
    if col_filters:
        filtered_dax = wrap_dax_with_filters(filtered_dax, col_filters, pattern)
    if meas_filters:
        filtered_dax = wrap_dax_with_having(filtered_dax, meas_filters)

    # Also wrap pivot query with filters if applicable
    if pivot_dax_query and active_filters:
        if col_filters:
            pivot_dax_query = wrap_dax_with_filters(
                f"EVALUATE\n{pivot_dax_query.split('EVALUATE')[1].strip()}" if "EVALUATE" in pivot_dax_query else pivot_dax_query,
                col_filters, "Pattern 3M: Matrix Pivot"
            )
            # Re-extract just the pivot part (wrap_dax_with_filters already adds EVALUATE)
            pivot_dax_query = pivot_dax_query

    # Wrap with post-aggregation conditions (FILTER) from explicit having
    if having_exprs:
        filtered_dax = wrap_dax_with_having(filtered_dax, having_exprs)
        having_applied = "; ".join(having_exprs)
    else:
        having_applied = ""

    filters_applied = "; ".join(active_filters) if active_filters else ""
    preset_applied = "; ".join(preset_filters) if preset_filters else ""

    return {
        "page": page,
        "visual_name": visual_name,
        "visual_type": visual_type,
        "pattern": pattern,
        "dax_query": filtered_dax,
        "base_dax_query": base_dax,
        "filters_applied": filters_applied,
        "having_applied": having_applied,
        "preset_filters_applied": preset_applied,
        "matrix_columns": matrix_col_info,
        "values_query": values_query,
        "pivot_dax_query": pivot_dax_query,
        "auto_flat_measures": auto_flat_measures,
    }


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Generate DAX queries from PBI metadata extractor output."
    )
    parser.add_argument("input_excel", help="Path to the metadata extractor Excel file")
    parser.add_argument("output_excel", nargs="?", default=None,
                        help="Path for output Excel file (default: dax_queries_<input_name>.xlsx)")
    parser.add_argument("--visual", default=None,
                        help="Extract DAX for a single visual (partial name match). "
                             "Prints the query to stdout instead of generating a full Excel.")
    parser.add_argument("--filter", action="append", default=None, dest="filters",
                        help="Pre-aggregation DAX filter expression — CALCULATETABLE (can be repeated). "
                             "E.g.: --filter \"'Store'[Store Type] = \\\"New Store\\\"\"")
    parser.add_argument("--having", action="append", default=None, dest="having",
                        help="Post-aggregation DAX condition — FILTER on results (can be repeated). "
                             "E.g.: --having \"[TotalSales] > 1000000\"")
    parser.add_argument("--model-root", default=None,
                        help="Path to semantic model definition (fallback for measure formula lookup "
                             "when formulas aren't in the metadata Excel)")
    args = parser.parse_args()

    if not os.path.exists(args.input_excel):
        print(f"Error: File not found: {args.input_excel}")
        sys.exit(1)

    # Load semantic model if --model-root provided (fallback for measure formulas)
    model = None
    if args.model_root:
        try:
            from tmdl_parser import parse_semantic_model
        except ImportError:
            # When run from project root, try skills/ relative import
            sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
            from tmdl_parser import parse_semantic_model
        model = parse_semantic_model(args.model_root)
        print(f"Loaded semantic model from: {args.model_root} "
              f"({len(model.measures)} measures)")

    visuals, page_filters, bookmarks, filter_expr_data = read_extractor_output(args.input_excel)

    # --- Single visual mode ---
    if args.visual:
        result = get_single_visual_query(visuals, page_filters, args.visual,
                                         args.filters, args.having, model,
                                         filter_expr_data)
        if result:
            print(f"Page: {result['page']}")
            print(f"Visual: {result['visual_name']} ({result['visual_type']})")
            print(f"Pattern: {result['pattern']}")
            if result['preset_filters_applied']:
                print(f"Preset filters: {result['preset_filters_applied']}")
            if result['filters_applied']:
                print(f"All filters: {result['filters_applied']}")
            if result['having_applied']:
                print(f"Having: {result['having_applied']}")
            if result.get('matrix_columns'):
                mc_str = ", ".join(
                    f"'{mc['table']}'[{mc['column']}]" for mc in result['matrix_columns']
                )
                print(f"\nMatrix column-axis field: {mc_str}")
                print(f"Preflight VALUES query:")
                print(result['values_query'])
            if result.get('pivot_dax_query'):
                print(f"\nPivoted DAX:")
                print(result['pivot_dax_query'])
            print(f"\nFiltered DAX:")
            print(result['dax_query'])
            print(f"\nBase DAX:")
            print(result['base_dax_query'])
        return

    # --- Full output mode ---
    # Default output name
    if args.output_excel is None:
        base = os.path.splitext(os.path.basename(args.input_excel))[0]
        args.output_excel = f"dax_queries_{base}.xlsx"

    print(f"Reading: {args.input_excel}")
    print(f"Found {len(visuals)} visuals across {len(set(k[0] for k in visuals))} pages")
    if page_filters:
        print(f"Found page-level filters on: {', '.join(page_filters.keys())}")
    if filter_expr_data:
        print(f"Found {len(filter_expr_data)} filter expressions — will generate Filtered DAX column")

    # Build bookmark DAX queries if bookmarks are present
    bookmark_queries = []
    if bookmarks:
        print(f"Found {len(bookmarks)} bookmark rows — generating bookmark DAX queries")
        bookmark_queries = build_bookmark_queries(bookmarks, visuals, page_filters, model)

    count = write_output(visuals, page_filters, args.output_excel, bookmark_queries,
                         filter_expr_data, model)

    print(f"\nGenerated DAX queries for {count} visuals")
    print(f"Output: {args.output_excel}")

    # Print summary
    print("\n--- Summary ---")
    for (page, _key), data in visuals.items():
        visual_name = data["visual_name"]
        grouping, measures, filters, slicer_fields, matrix_columns = classify_visual_fields(data["fields"])
        if page in page_filters:
            filters.extend(page_filters[page])
        pattern, _ = build_dax_query(grouping, measures, filters, slicer_fields,
                                     data["visual_type"], model, matrix_columns=matrix_columns)
        filter_note = f" [has filters]" if filters else ""
        print(f"  {page} / {visual_name} ({data['visual_type']}) -> {pattern}{filter_note}")

    if bookmark_queries:
        bm_names = sorted(set(bq["bookmark_name"] for bq in bookmark_queries))
        print(f"\n--- Bookmark DAX Queries ---")
        for bm_name in bm_names:
            bm_vis = [bq for bq in bookmark_queries if bq["bookmark_name"] == bm_name]
            print(f"  {bm_name}: {len(bm_vis)} visible visuals with DAX queries")


if __name__ == "__main__":
    main()
